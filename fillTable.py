# coding=utf-8
import os
from openpyxl import load_workbook
import pdfOcr
import removeSpacesAndNewlines
import ollamaAns

def fillTable(path,ollamaServer,modelName):
    # 找到path下的excel文件
    excel_file = None 
    for file in os.listdir(path):
        if file.endswith(".xlsx") or file.endswith(".xls"):
            excel_file = os.path.join(path, file)
            break

    if not excel_file:
        print("Excel文件未找到")
        return

    # 读取excel文件
    workbook = load_workbook(excel_file)
    sheet = workbook.active 

    # 获取excel表格的最大列数，并在最后一列添加'文件名'标题
    # 注意空表的max_row为1，所以要判断是否为空表
    if sheet.max_row  == 1 and sheet.max_column  == 1:
        print("Excel文件为空，开始提取文件名。")
        colmax = 1
        cell = sheet.cell(1, colmax)
        cell.value = '文件名'         
        # 遍历path下除excel文件以外的每个文件
        other_files = [f for f in os.listdir(path) if not f.endswith(".xlsx") and not f.endswith(".xls")]
        for i, file in enumerate(other_files):                                    
            cell = sheet.cell(i + 2, colmax)
            cell.value = file
        workbook.save(excel_file)

    #如果不是空表，就填写表格
    else:
        print("非空表，开始填写。")
        colmax = sheet.max_column + 1
        cell = sheet.cell(1, colmax)
        cell.value = '信息来源' 
        # 遍历path下除excel文件以外的每个文件
        other_files = [f for f in os.listdir(path) if not f.endswith(".xlsx") and not f.endswith(".xls")]
        for i, file in enumerate(other_files):
            # 获取文件路径
            pathFile = os.path.join(path, file)
            # 调用pdf_ocr函数获取文件的文本
            text0 = pdfOcr.pdfOcr(pathFile)
            # 调用remove_spaces_and_newlines函数处理文本        
            text0 = removeSpacesAndNewlines.removeSpacesAndNewlines(text0)
            # 调用remove_special_characters函数处理文本         

            # 遍历Excel文件中的每一列，填写第i+1行的空白单元格
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=i + 2, column=col)
                # 只处理空白单元格
                if cell.value is None:
                    # 获取该列第一行的内容
                    header_text = sheet.cell(row=1, column=col).value
                    # 调用ollamaAns函数，获取返回值并填入空白单元格
                    if header_text:
                        cell_value = ollamaAns.ollamaAns(header_text, text0,ollama_url=ollamaServer,model_name=modelName)
                        cell.value = cell_value.json()['response']
                        print(cell.value,"filled")                         
            cell = sheet.cell(i + 2, colmax)
            cell.value = file
            workbook.save(excel_file)  
    # 保存修改后的excel文件
    workbook.save(excel_file)
    print("已全部处理完成，保存文件：", excel_file)

if __name__ == "__main__":
    ollamaServer = os.getenv("OllamaServer", "http://10.8.0.3:11434")  # 设置默认值
    modelName = os.getenv("ModelName", "qwen2.5:7b")  # 设置默认值
    # fillTable("/app/data",ollamaServer,modelName)
    fillTable(r".\data",ollamaServer,modelName)